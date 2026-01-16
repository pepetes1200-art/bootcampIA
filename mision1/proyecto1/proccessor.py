import re
from openpyxl import load_workbook
#========================================
# funcion celan_id
#  PUNTO elimina caracteres no numericos de un documento
#"cc75.88.56" = "758856"
# =======================================
def clean_id(value):
    # elimina carateres no numeros de un docummento
    if value is None:
        return""
    return re.sub(r"\D",'',str(value))
#========================================
#  funcion merge_name
#  une nombree y apellido en un solo campo
#====================================
def merge_name (name, lastname):
    if name is None:
        name=""
    if lastname is None:
        lastname=""
    return f"{name} {lastname}".strip()

def process_excel(path):
    # Acceso ala hoja de llamada"datos"
    wb= load_workbook(path)
    ws = wb["Datos"]
    # recorrer todas las filas desde la fila 2 
    for row in range (2,ws.max_row+1):
        #columnas D: identificador limpio
        ws[f"D{row}"] =clean_id(ws[f"A{row}"].value)
        # columna E: nombre completo 
        ws[f"E{row}"]=merge_name(
        ws[f"B{row}"].value,
        ws[f"C{row}"].value
        )
        
        